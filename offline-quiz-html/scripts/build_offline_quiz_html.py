#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Iterable, Sequence

import openpyxl

DEFAULT_OUTPUT_NAME = "离线题库刷题.html"
DEFAULT_ANSWER_SUFFIX = ".answers.md"
SUPPORTED_MARKDOWN_SUFFIXES = {".md", ".markdown", ".txt"}
SUPPORTED_XLSX_SUFFIXES = {".xlsx", ".xlsm"}
APP_STATE_KEY = "offline_quiz_html_state_v1"

SOURCE_FILE_HEADER = "来源文件"
QUESTION_STEM_HEADER = "题目"
QUESTION_TYPE_SINGLE = "单选题"
QUESTION_TYPE_MULTI = "多选题"
QUESTION_TYPE_JUDGEMENT = "判断题"
QUESTION_TYPE_FILL = "填空题"
QUESTION_TYPE_SHORT = "问答题"

HEADING_RE = re.compile(r"^\s*#{1,6}\s+(.+?)\s*$")
QUESTION_RE = re.compile(r"^\s*(\d+)[\.、]\s*(.+?)\s*$")
OPTION_RE = re.compile(r"^\s*([A-F])\s*[\.、:：]\s*(.+?)\s*$")
ANSWER_LINE_RE = re.compile(r"^\s*(?:答案|参考答案|正确答案)\s*[：:]\s*(.+?)\s*$")
ANALYSIS_LINE_RE = re.compile(r"^\s*(?:解析|说明|备注)\s*[：:]\s*(.+?)\s*$")
TRAILING_PARENTHESES_RE = re.compile(r"^(.*?)[(（]([^()（）]+)[)）]\s*$")
INLINE_ANSWER_RE = re.compile(r"[【\[]\s*(?:答案|参考答案|正确答案)\s*[：:]\s*([^】\]]+)\s*[】\]]")


@dataclass(frozen=True)
class Question:
    uid: str
    source_kind: str
    group_id: str
    group_name: str
    source_label: str
    group_note: str
    question_number: str
    question_type: str
    response_mode: str
    stem: str
    options: dict[str, str]
    answer: str
    analysis: str


@dataclass(frozen=True)
class GroupInfo:
    id: str
    name: str
    source_label: str
    note: str
    count: int


@dataclass
class _MarkdownGroupBuilder:
    group_name: str
    group_note_lines: list[str]
    questions: list[Question]


@dataclass
class _MarkdownQuestionBuilder:
    number: str
    stem_lines: list[str]
    option_lines: list[str]
    answer_lines: list[str]
    analysis_lines: list[str]


def _normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\u3000", " ").strip()


def _normalize_choice_answer(value: object) -> str:
    text = _normalize_text(value).upper()
    if not text:
        return ""
    if text in {"对", "正确", "TRUE", "T", "YES", "Y"}:
        return "A"
    if text in {"错", "错误", "FALSE", "F", "NO", "N"}:
        return "B"
    letters = [ch for ch in text if ch in "ABCDEF"]
    if not letters:
        return text
    ordered: list[str] = []
    for ch in letters:
        if ch not in ordered:
            ordered.append(ch)
    return "".join(sorted(ordered)) if len(ordered) > 1 else ordered[0]


def _split_answer_variants(value: str) -> list[str]:
    parts = [part.strip() for part in re.split(r"[|/;；／]", value) if part.strip()]
    return parts or ([value.strip()] if value.strip() else [])


def _normalize_free_text(value: str) -> str:
    return re.sub(r"[\s\u3000]+", "", value).lower()


def _compare_answer(selected: str, answer: str, response_mode: str) -> str:
    if not selected:
        return "未答"
    if response_mode == "text":
        selected_norm = _normalize_free_text(selected)
        if not answer.strip():
            return "错误"
        for variant in _split_answer_variants(answer):
            if _normalize_free_text(variant) == selected_norm:
                return "正确"
        return "错误"
    return "正确" if _normalize_choice_answer(selected) == _normalize_choice_answer(answer) else "错误"


def _question_uid(source_name: str, group_id: str, stable_id: str) -> str:
    return f"{source_name}::{group_id}::{stable_id}"


def _collect_inputs(inputs: Iterable[str], recursive: bool = True) -> list[Path]:
    files: list[Path] = []
    seen: set[Path] = set()
    for raw in inputs:
        path = Path(raw)
        if path.is_file():
            if path.name.endswith(DEFAULT_ANSWER_SUFFIX):
                continue
            suffix = path.suffix.lower()
            if suffix in SUPPORTED_MARKDOWN_SUFFIXES or suffix in SUPPORTED_XLSX_SUFFIXES:
                resolved = path.resolve()
                if resolved not in seen:
                    seen.add(resolved)
                    files.append(path)
            continue
        if not path.is_dir():
            continue
        iterator = path.rglob("*") if recursive else path.iterdir()
        for candidate in sorted(iterator):
            if not candidate.is_file() or candidate.name.endswith(DEFAULT_ANSWER_SUFFIX):
                continue
            suffix = candidate.suffix.lower()
            if suffix not in SUPPORTED_MARKDOWN_SUFFIXES and suffix not in SUPPORTED_XLSX_SUFFIXES:
                continue
            resolved = candidate.resolve()
            if resolved in seen:
                continue
            seen.add(resolved)
            files.append(candidate)
    return sorted(files, key=lambda item: str(item).lower())


def _load_xlsx_questions(paths: Sequence[Path]) -> list[Question]:
    questions: list[Question] = []
    for path in paths:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            for ws in wb.worksheets:
                header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                if not header or len(header) < 15:
                    continue
                if _normalize_text(header[1]) != SOURCE_FILE_HEADER or _normalize_text(header[4]) != QUESTION_STEM_HEADER:
                    continue
                for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if len(row) < 5:
                        continue
                    source_file = _normalize_text(row[1])
                    stem = _normalize_text(row[4])
                    if not source_file or not stem:
                        continue
                    options = {
                        label: _normalize_text(row[idx])
                        for idx, label in enumerate("ABCDEF", start=5)
                        if idx < len(row) and _normalize_text(row[idx])
                    }
                    answer = _normalize_choice_answer(row[12] if len(row) > 12 else "")
                    analysis = _normalize_text(row[14] if len(row) > 14 else "")
                    if not analysis and len(row) > 13:
                        analysis = _normalize_text(row[13])
                    qtype = _normalize_text(row[3]) or QUESTION_TYPE_SINGLE
                    group_id = f"{path.name}::{source_file}"
                    questions.append(
                        Question(
                            uid=_question_uid(path.name, group_id, f"{ws.title}:{row_index}"),
                            source_kind="xlsx",
                            group_id=group_id,
                            group_name=source_file,
                            source_label=path.name,
                            group_note="",
                            question_number=_normalize_text(row[2]),
                            question_type=qtype,
                            response_mode="choice",
                            stem=stem,
                            options=options,
                            answer=answer,
                            analysis=analysis,
                        )
                    )
        finally:
            wb.close()
    return questions


def _extract_inline_answer(text: str) -> tuple[str, str]:
    cleaned = _normalize_text(text)
    if not cleaned:
        return "", ""
    match = TRAILING_PARENTHESES_RE.match(cleaned)
    if match:
        stem = _normalize_text(match.group(1))
        answer = _normalize_text(match.group(2))
        if stem and answer:
            return stem, answer
    inline = INLINE_ANSWER_RE.search(cleaned)
    if inline:
        answer = _normalize_text(inline.group(1))
        stem = INLINE_ANSWER_RE.sub("", cleaned).strip()
        return stem, answer
    return cleaned, ""


def _new_markdown_group(group_name: str) -> _MarkdownGroupBuilder:
    return _MarkdownGroupBuilder(group_name=group_name, group_note_lines=[], questions=[])


def _finalize_markdown_question(
    source_name: str,
    group: _MarkdownGroupBuilder,
    question_builder: _MarkdownQuestionBuilder | None,
    question_index: int,
) -> tuple[_MarkdownQuestionBuilder | None, int]:
    if question_builder is None:
        return None, question_index

    stem = "\n".join(line for line in question_builder.stem_lines if _normalize_text(line)).strip()
    if not stem and question_builder.option_lines:
        stem = "\n".join(question_builder.option_lines).strip()
    answer = "\n".join(line for line in question_builder.answer_lines if _normalize_text(line)).strip()
    analysis = "\n".join(line for line in question_builder.analysis_lines if _normalize_text(line)).strip()
    options: dict[str, str] = {}
    for line in question_builder.option_lines:
        option_match = OPTION_RE.match(line)
        if option_match:
            options[option_match.group(1)] = _normalize_text(option_match.group(2))

    response_mode = "choice" if options else "text"
    if response_mode == "text" and not answer:
        stem, answer = _extract_inline_answer(stem)
    elif response_mode == "choice" and not answer:
        _, extracted = _extract_inline_answer(stem)
        if extracted:
            answer = extracted

    if response_mode == "choice" and len(_normalize_choice_answer(answer)) > 1:
        question_type = QUESTION_TYPE_MULTI
    elif response_mode == "choice" and _normalize_choice_answer(answer) in {"A", "B"} and any(
        _normalize_text(opt) for opt in options.values()
    ):
        question_type = QUESTION_TYPE_SINGLE
    else:
        question_type = QUESTION_TYPE_FILL if response_mode == "text" else QUESTION_TYPE_SINGLE

    question = Question(
        uid=_question_uid(source_name, group.group_name, f"md-{question_index}"),
        source_kind="markdown",
        group_id=f"{source_name}::{group.group_name}",
        group_name=group.group_name,
        source_label=source_name,
        group_note="\n".join(line for line in group.group_note_lines if _normalize_text(line)).strip(),
        question_number=question_builder.number,
        question_type=question_type,
        response_mode=response_mode,
        stem=stem,
        options=options,
        answer=answer,
        analysis=analysis,
    )
    group.questions.append(question)
    return None, question_index + 1


def _load_markdown_questions(paths: Sequence[Path]) -> list[Question]:
    questions: list[Question] = []
    for path in paths:
        lines = path.read_text(encoding="utf-8").splitlines()
        source_name = path.name
        current_group = _new_markdown_group(path.stem)
        current_question: _MarkdownQuestionBuilder | None = None
        question_index = 1
        seen_question = False
        heading_index = 0

        def flush_question() -> None:
            nonlocal current_question, question_index
            current_question, question_index = _finalize_markdown_question(source_name, current_group, current_question, question_index)

        def flush_group() -> None:
            nonlocal current_group, current_question, question_index, seen_question
            flush_question()
            questions.extend(current_group.questions)
            current_group = _new_markdown_group(path.stem)
            current_question = None
            question_index = 1
            seen_question = False

        for raw_line in lines:
            line = raw_line.lstrip("\ufeff").rstrip()
            heading_match = HEADING_RE.match(line)
            if heading_match:
                if current_group.questions or current_group.group_note_lines or seen_question:
                    flush_group()
                heading_index += 1
                current_group.group_name = _normalize_text(heading_match.group(1)) or f"{path.stem}-{heading_index}"
                continue

            question_match = QUESTION_RE.match(line)
            if question_match:
                flush_question()
                seen_question = True
                current_question = _MarkdownQuestionBuilder(
                    number=_normalize_text(question_match.group(1)),
                    stem_lines=[_normalize_text(question_match.group(2))],
                    option_lines=[],
                    answer_lines=[],
                    analysis_lines=[],
                )
                continue

            if current_question is None:
                if _normalize_text(line):
                    current_group.group_note_lines.append(line)
                elif current_group.group_note_lines and current_group.group_note_lines[-1] != "":
                    current_group.group_note_lines.append("")
                continue

            answer_match = ANSWER_LINE_RE.match(line)
            if answer_match:
                current_question.answer_lines.append(answer_match.group(1))
                continue
            analysis_match = ANALYSIS_LINE_RE.match(line)
            if analysis_match:
                current_question.analysis_lines.append(analysis_match.group(1))
                continue
            option_match = OPTION_RE.match(line)
            if option_match:
                current_question.option_lines.append(line)
                continue
            if _normalize_text(line):
                current_question.stem_lines.append(_normalize_text(line))
            elif current_question.stem_lines and current_question.stem_lines[-1] != "":
                current_question.stem_lines.append("")

        flush_group()
    return questions


def load_questions(paths: Sequence[Path]) -> list[Question]:
    questions: list[Question] = []
    xlsx_paths = [path for path in paths if path.suffix.lower() in SUPPORTED_XLSX_SUFFIXES]
    markdown_paths = [path for path in paths if path.suffix.lower() in SUPPORTED_MARKDOWN_SUFFIXES]
    if xlsx_paths:
        questions.extend(_load_xlsx_questions(xlsx_paths))
    if markdown_paths:
        questions.extend(_load_markdown_questions(markdown_paths))
    return questions


def _group_questions(questions: Sequence[Question]) -> list[tuple[GroupInfo, list[Question]]]:
    grouped: dict[str, list[Question]] = defaultdict(list)
    for question in questions:
        grouped[question.group_id].append(question)
    ordered: list[tuple[GroupInfo, list[Question]]] = []
    for group_id, items in grouped.items():
        first = items[0]
        ordered.append(
            (
                GroupInfo(
                    id=group_id,
                    name=first.group_name,
                    source_label=first.source_label,
                    note=first.group_note,
                    count=len(items),
                ),
                list(items),
            )
        )
    return ordered


def _json_script(value: object) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":")).replace("</", "<\\/")


def render_answer_markdown(questions: Sequence[Question]) -> str:
    grouped = _group_questions(questions)
    lines: list[str] = ["# Offline Quiz Answers", ""]
    for group, items in grouped:
        lines.append(f"## {group.name}")
        if group.source_label:
            lines.append(f"- Source: {group.source_label}")
        if group.note:
            lines.append("")
            lines.append(group.note)
        lines.append("")
        for question in items:
            label = question.question_number or "?"
            lines.append(f"### {label}. {question.stem}")
            for key, text in question.options.items():
                lines.append(f"- {key}. {text}")
            lines.append(f"- Answer: {question.answer or 'N/A'}")
            if question.analysis:
                lines.append(f"- Analysis: {question.analysis}")
            lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def render_html(questions: Sequence[Question]) -> str:
    grouped = _group_questions(questions)
    group_payload = [asdict(group) for group, _ in grouped]
    question_payload = [asdict(question) for question in questions]
    template = r"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>__TITLE__</title>
  <style>
    :root {
      color-scheme: light;
      --bg: #f4efe6;
      --bg-soft: #fffaf1;
      --ink: #1f1b16;
      --muted: #6b6258;
      --brand: #c06a2b;
      --brand-2: #0f6b5b;
      --line: rgba(31, 27, 22, 0.12);
      --shadow: 0 18px 50px rgba(48, 28, 11, 0.12);
      --ok: #157a50;
      --bad: #b53f2d;
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      min-height: 100vh;
      color: var(--ink);
      background:
        radial-gradient(circle at top left, rgba(192, 106, 43, 0.12), transparent 28%),
        radial-gradient(circle at top right, rgba(15, 107, 91, 0.12), transparent 24%),
        linear-gradient(180deg, #f8f2e7 0%, #f0e5d3 100%);
      font-family: "Noto Serif SC", "Source Han Serif SC", "Songti SC", "Microsoft YaHei", serif;
    }
    .shell { max-width: 1500px; margin: 0 auto; padding: 22px; }
    .hero {
      display: grid;
      grid-template-columns: 1.25fr 0.95fr;
      gap: 18px;
      margin-bottom: 18px;
    }
    .panel {
      background: rgba(255, 250, 241, 0.92);
      border: 1px solid var(--line);
      border-radius: 24px;
      box-shadow: var(--shadow);
      backdrop-filter: blur(8px);
    }
    .hero-main { padding: 24px; position: relative; overflow: hidden; }
    .hero-main::after {
      content: '';
      position: absolute;
      inset: auto -10% -40% auto;
      width: 320px;
      height: 320px;
      background: radial-gradient(circle, rgba(192, 106, 43, 0.18), transparent 65%);
      pointer-events: none;
    }
    .eyebrow { text-transform: uppercase; letter-spacing: .22em; color: var(--brand-2); font-size: 12px; font-weight: 700; }
    h1 { margin: 12px 0 10px; font-size: clamp(32px, 4vw, 56px); line-height: 1.02; }
    .hero-copy { max-width: 68ch; color: var(--muted); font-size: 16px; line-height: 1.75; }
    .stats { display: grid; grid-template-columns: repeat(3, minmax(0, 1fr)); gap: 12px; margin-top: 18px; }
    .stat { background: rgba(255, 255, 255, 0.72); border: 1px solid var(--line); border-radius: 18px; padding: 14px 16px; }
    .stat strong { display: block; font-size: 28px; margin-bottom: 4px; }
    .hero-side { padding: 20px; display: flex; flex-direction: column; gap: 14px; justify-content: space-between; }
    .mode-buttons { display: grid; grid-template-columns: repeat(2, minmax(0, 1fr)); gap: 12px; }
    .mode-btn, .action-btn, .group-btn, .question-btn, .mini-btn {
      border: 0;
      border-radius: 16px;
      padding: 13px 16px;
      cursor: pointer;
      font: inherit;
      transition: transform .18s ease, box-shadow .18s ease, background .18s ease;
    }
    .mode-btn { background: linear-gradient(135deg, var(--brand), #ea9151); color: #fff; font-weight: 700; box-shadow: 0 12px 24px rgba(192, 106, 43, 0.22); }
    .mode-btn.secondary { background: linear-gradient(135deg, var(--brand-2), #2b8c7b); }
    .action-btn, .mini-btn { background: #fff; border: 1px solid var(--line); color: var(--ink); }
    .mode-btn:hover, .action-btn:hover, .group-btn:hover, .question-btn:hover, .mini-btn:hover { transform: translateY(-1px); }
    .layout { display: grid; grid-template-columns: 320px minmax(0, 1fr); gap: 18px; align-items: start; }
    .sidebar { padding: 18px; position: sticky; top: 18px; }
    .sidebar h2, .content h2 { margin: 0 0 12px; font-size: 22px; }
    .group-list { display: grid; gap: 10px; max-height: calc(100vh - 260px); overflow: auto; padding-right: 4px; }
    .group-btn { width: 100%; text-align: left; background: #fff; border: 1px solid var(--line); }
    .group-btn.active { border-color: rgba(15, 107, 91, 0.38); box-shadow: inset 0 0 0 1px rgba(15, 107, 91, 0.22); background: linear-gradient(135deg, rgba(15, 107, 91, 0.08), rgba(192, 106, 43, 0.06)); }
    .group-title { display: flex; justify-content: space-between; gap: 10px; font-size: 15px; font-weight: 700; }
    .group-meta { color: var(--muted); font-size: 13px; margin-top: 6px; line-height: 1.5; }
    .content { padding: 18px; }
    .toolbar { display: flex; flex-wrap: wrap; gap: 10px; margin-bottom: 14px; }
    .toolbar .primary { background: linear-gradient(135deg, var(--brand-2), #2b8c7b); color: #fff; border: 0; }
    .group-card {
      border: 1px solid var(--line);
      border-radius: 22px;
      padding: 18px;
      background: rgba(255,255,255,.82);
      box-shadow: 0 14px 32px rgba(40, 23, 8, 0.08);
      margin-bottom: 16px;
    }
    .group-card h3 { margin: 0; font-size: 22px; }
    .group-top { display: flex; justify-content: space-between; gap: 12px; align-items: start; margin-bottom: 12px; }
    .group-note {
      margin: 12px 0 18px;
      padding: 14px 16px;
      background: rgba(15, 107, 91, 0.06);
      border: 1px dashed rgba(15, 107, 91, 0.22);
      border-radius: 16px;
      color: var(--muted);
      white-space: pre-wrap;
      line-height: 1.7;
    }
    .group-actions { display: flex; flex-wrap: wrap; gap: 10px; }
    .question-list { display: grid; gap: 14px; margin-top: 16px; }
    .question-card { border: 1px solid var(--line); border-radius: 20px; padding: 18px; background: rgba(255,255,255,.88); }
    .question-head { display: flex; justify-content: space-between; gap: 12px; align-items: start; margin-bottom: 10px; }
    .badges { display: flex; flex-wrap: wrap; gap: 8px; }
    .badge { display: inline-flex; align-items: center; gap: 6px; padding: 6px 10px; border-radius: 999px; font-size: 12px; color: var(--muted); background: rgba(31, 27, 22, 0.05); }
    .stem { font-size: 17px; line-height: 1.8; white-space: pre-wrap; }
    .options { display: grid; gap: 10px; margin-top: 14px; }
    .option {
      display: grid;
      grid-template-columns: auto auto 1fr;
      gap: 10px;
      align-items: start;
      padding: 12px 14px;
      border: 1px solid rgba(31, 27, 22, 0.1);
      border-radius: 14px;
      background: rgba(255, 255, 255, 0.85);
      cursor: pointer;
    }
    .option-key {
      min-width: 24px;
      color: var(--brand-2);
      font-weight: 800;
    }
    .option-text { line-height: 1.7; white-space: pre-wrap; }
    .text-answer {
      width: 100%;
      margin-top: 14px;
      border: 1px solid rgba(31, 27, 22, 0.14);
      border-radius: 14px;
      padding: 14px 16px;
      font: inherit;
      background: #fff;
      color: var(--ink);
      outline: none;
    }
    .text-answer:focus { border-color: rgba(15, 107, 91, 0.42); box-shadow: 0 0 0 4px rgba(15, 107, 91, 0.08); }
    .question-actions { display: flex; flex-wrap: wrap; gap: 10px; margin-top: 14px; }
    .status {
      margin-top: 12px;
      padding: 12px 14px;
      border-radius: 14px;
      background: rgba(31, 27, 22, 0.04);
      color: var(--muted);
      line-height: 1.7;
      white-space: pre-wrap;
    }
    .status.ok { background: rgba(21, 122, 80, 0.1); color: var(--ok); }
    .status.bad { background: rgba(181, 63, 45, 0.1); color: var(--bad); }
    .empty { padding: 28px; text-align: center; color: var(--muted); }
    @media (max-width: 1180px) {
      .hero, .layout { grid-template-columns: 1fr; }
      .sidebar { position: static; }
      .group-list { max-height: none; }
    }
  </style>
</head>
<body>
  <div class="shell">
    <div class="hero">
      <section class="panel hero-main">
        <div class="eyebrow">Offline quiz builder</div>
        <h1>__TITLE__</h1>
        <div class="hero-copy">__HERO_COPY__</div>
        <div class="stats">
          <div class="stat"><strong id="stat-total">0</strong><span>Questions</span></div>
          <div class="stat"><strong id="stat-groups">0</strong><span>Groups</span></div>
          <div class="stat"><strong id="stat-correct">0</strong><span>Correct</span></div>
        </div>
      </section>
      <aside class="panel hero-side">
        <div>
          <h2>Actions</h2>
          <div class="mode-buttons">
            <button class="mode-btn primary" id="grade-all">Submit all</button>
            <button class="mode-btn secondary" id="reset-all">Reset all</button>
          </div>
          <div class="toolbar" style="margin-top:12px;">
            <button class="action-btn primary" id="jump-first">Jump first group</button>
            <button class="action-btn" id="clear-highlight">Clear highlights</button>
          </div>
        </div>
        <div>
          <div class="badge" id="summary-line">Ready.</div>
        </div>
      </aside>
    </div>

    <div class="layout">
      <aside class="panel sidebar">
        <h2>Groups</h2>
        <div class="group-list" id="group-list"></div>
      </aside>
      <main class="panel content">
        <h2>Practice</h2>
        <div class="toolbar">
          <button class="action-btn primary" id="expand-all">Expand all groups</button>
          <button class="action-btn" id="collapse-all">Collapse all groups</button>
        </div>
        <div id="group-area"></div>
      </main>
    </div>
  </div>

  <script>
    const APP_STATE_KEY = __APP_STATE_KEY__;
    const QUESTIONS = __QUESTION_DATA__;
    const GROUPS = __GROUP_DATA__;
    const TYPE_MULTI = "多选题";
    const TYPE_JUDGEMENT = "判断题";

    function escapeHtml(text) {
      return String(text)
        .replace(/&/g, '&amp;')
        .replace(/</g, '&lt;')
        .replace(/>/g, '&gt;')
        .replace(/"/g, '&quot;')
        .replace(/'/g, '&#39;');
    }

    function normalizeChoice(answer) {
      if (!answer) return '';
      const text = String(answer).toUpperCase();
      if (['对', '正确', 'TRUE', 'T', 'YES', 'Y'].includes(text)) return 'A';
      if (['错', '错误', 'FALSE', 'F', 'NO', 'N'].includes(text)) return 'B';
      const letters = Array.from(new Set(text.split('').filter(ch => 'ABCDEF'.includes(ch)))).sort();
      return letters.join('') || text;
    }

    function normalizeText(answer) {
      return String(answer || '').replace(/[\s\u3000]+/g, '').toLowerCase();
    }

    function splitVariants(answer) {
      const raw = String(answer || '').trim();
      if (!raw) return [];
      return raw.split(/[|/;；／]/).map(item => item.trim()).filter(Boolean);
    }

    function compareTextAnswer(selected, answer) {
      const selectedNorm = normalizeText(selected);
      if (!selectedNorm) return '未答';
      const variants = splitVariants(answer);
      if (!variants.length) return '错误';
      return variants.some(variant => normalizeText(variant) === selectedNorm) ? '正确' : '错误';
    }

    function compareAnswer(question, selected) {
      if (!selected) return '未答';
      if (question.response_mode === 'text') {
        return compareTextAnswer(selected, question.answer);
      }
      return normalizeChoice(selected) === normalizeChoice(question.answer) ? '正确' : '错误';
    }

    function prettyAnswer(question) {
      if (question.response_mode === 'text') {
        return question.answer || '未提供';
      }
      if (question.question_type === TYPE_JUDGEMENT) {
        return normalizeChoice(question.answer) === 'A' ? '正确' : '错误';
      }
      return question.answer || '未提供';
    }

    function questionInputValue(questionCard) {
      const textInput = questionCard.querySelector('[data-text-input]');
      if (textInput) return textInput.value.trim();
      return Array.from(questionCard.querySelectorAll('input')).filter(input => input.checked).map(input => input.value).join('');
    }

    function renderQuestion(question, index) {
      const wrap = document.createElement('article');
      wrap.className = 'question-card';
      wrap.dataset.questionId = question.uid;
      wrap.dataset.groupId = question.group_id;
      const options = Object.entries(question.options || {});
      const optionType = question.question_type === TYPE_MULTI ? 'checkbox' : 'radio';
      const title = question.question_number ? `${question.question_number}.` : `Q${index + 1}`;
      wrap.innerHTML = `
        <div class="question-head">
          <div>
            <div class="badges">
              <span class="badge">${escapeHtml(title)}</span>
              <span class="badge">${escapeHtml(question.question_type || '题目')}</span>
              <span class="badge">${escapeHtml(question.source_label || '')}</span>
            </div>
          </div>
          <span class="badge">${escapeHtml(question.response_mode === 'text' ? 'Text' : 'Choice')}</span>
        </div>
        <div class="stem">${escapeHtml(question.stem || '')}</div>
        <div class="options" data-options></div>
        <input class="text-answer" data-text-input placeholder="请直接输入答案" style="display:none;">
        <div class="question-actions">
          <button type="button" class="question-btn primary" data-submit>Submit</button>
          <button type="button" class="question-btn" data-reset>Reset</button>
        </div>
        <div class="status" data-status>先作答，再提交查看判定结果、正确答案和解析。</div>
      `;

      const optionsHost = wrap.querySelector('[data-options]');
      const textInput = wrap.querySelector('[data-text-input]');
      const submitButton = wrap.querySelector('[data-submit]');
      const resetButton = wrap.querySelector('[data-reset]');
      const status = wrap.querySelector('[data-status]');

      if (options.length && question.response_mode !== 'text') {
        optionsHost.style.display = 'grid';
        options.forEach(([label, text]) => {
          const row = document.createElement('label');
          row.className = 'option';
          row.innerHTML = `
            <input type="${optionType}" name="${question.uid}" value="${label}">
            <span class="option-key">${escapeHtml(label)}</span>
            <span class="option-text">${escapeHtml(text)}</span>
          `;
          optionsHost.appendChild(row);
        });
      } else {
        optionsHost.remove();
        textInput.style.display = 'block';
      }

      function renderPendingStatus() {
        const selected = questionInputValue(wrap);
        if (!selected) {
          status.className = 'status';
          status.textContent = '先作答，再提交查看判定结果、正确答案和解析。';
          return;
        }
        status.className = 'status';
        status.textContent = `已选择：${selected}`;
      }

      function renderResult() {
        const selected = questionInputValue(wrap);
        const verdict = compareAnswer(question, selected);
        status.className = 'status' + (verdict === '正确' ? ' ok' : verdict === '错误' ? ' bad' : '');
        const analysis = question.analysis ? question.analysis : '暂无解析';
        status.innerHTML = `<strong>${verdict}</strong> · 正确答案：<strong>${escapeHtml(prettyAnswer(question))}</strong><br>解析：${escapeHtml(analysis)}`;
        wrap.dataset.graded = '1';
        wrap.dataset.correct = verdict === '正确' ? '1' : '0';
        updateSummary();
        updateGroupState();
      }

      function resetQuestion() {
        wrap.querySelectorAll('input').forEach(input => {
          if (input.type === 'checkbox' || input.type === 'radio') input.checked = false;
          if (input.type === 'text') input.value = '';
        });
        delete wrap.dataset.graded;
        delete wrap.dataset.correct;
        renderPendingStatus();
        updateSummary();
        updateGroupState();
      }

      wrap.querySelectorAll('input').forEach(input => input.addEventListener('change', renderPendingStatus));
      if (textInput) textInput.addEventListener('input', renderPendingStatus);
      submitButton.addEventListener('click', renderResult);
      resetButton.addEventListener('click', resetQuestion);
      renderPendingStatus();
      return wrap;
    }

    function renderGroups() {
      const host = document.getElementById('group-area');
      const listHost = document.getElementById('group-list');
      host.innerHTML = '';
      listHost.innerHTML = '';
      GROUPS.forEach(group => {
        const items = QUESTIONS.filter(question => question.group_id === group.id);
        const listButton = document.createElement('button');
        listButton.className = 'group-btn';
        listButton.dataset.groupId = group.id;
        listButton.innerHTML = `
          <div class="group-title"><span>${escapeHtml(group.name)}</span><span>${items.length}</span></div>
          <div class="group-meta">${escapeHtml(group.source_label || '')}</div>
        `;
        listButton.addEventListener('click', () => {
          const target = document.querySelector(`[data-group-card="${CSS.escape(group.id)}"]`);
          if (target) target.scrollIntoView({behavior: 'smooth', block: 'start'});
        });
        listHost.appendChild(listButton);

        const section = document.createElement('section');
        section.className = 'group-card';
        section.dataset.groupCard = group.id;
        section.innerHTML = `
          <div class="group-top">
            <div>
              <h3>${escapeHtml(group.name)}</h3>
              <div class="badge">${escapeHtml(group.source_label || '')}</div>
            </div>
            <div class="group-actions">
              <button type="button" class="mini-btn" data-submit-group>Submit group</button>
              <button type="button" class="mini-btn" data-reset-group>Reset group</button>
              <button type="button" class="mini-btn" data-toggle-group>Collapse</button>
            </div>
          </div>
        `;
        if (group.note) {
          const note = document.createElement('div');
          note.className = 'group-note';
          note.textContent = group.note;
          section.appendChild(note);
        }
        const questionList = document.createElement('div');
        questionList.className = 'question-list';
        items.forEach((question, idx) => questionList.appendChild(renderQuestion(question, idx)));
        section.appendChild(questionList);

        const submitGroup = section.querySelector('[data-submit-group]');
        const resetGroup = section.querySelector('[data-reset-group]');
        const toggleGroup = section.querySelector('[data-toggle-group]');
        submitGroup.addEventListener('click', () => {
          questionList.querySelectorAll('.question-card [data-submit]').forEach(button => button.click());
        });
        resetGroup.addEventListener('click', () => {
          questionList.querySelectorAll('.question-card [data-reset]').forEach(button => button.click());
        });
        toggleGroup.addEventListener('click', () => {
          const hidden = section.dataset.collapsed === '1';
          section.dataset.collapsed = hidden ? '0' : '1';
          questionList.style.display = hidden ? 'grid' : 'none';
          const note = section.querySelector('.group-note');
          if (note) note.style.display = hidden ? 'block' : 'none';
          toggleGroup.textContent = hidden ? 'Collapse' : 'Expand';
          updateGroupState();
        });
        host.appendChild(section);
      });
      updateGroupState();
    }

    function updateSummary() {
      const total = QUESTIONS.length;
      const graded = document.querySelectorAll('.question-card[data-graded="1"]').length;
      const correct = document.querySelectorAll('.question-card[data-correct="1"]').length;
      document.getElementById('stat-total').textContent = total;
      document.getElementById('stat-groups').textContent = GROUPS.length;
      document.getElementById('stat-correct').textContent = correct;
      document.getElementById('summary-line').textContent = `Answered ${graded}/${total} · Correct ${correct}`;
    }

    function updateGroupState() {
      document.querySelectorAll('.group-btn').forEach(button => {
        const groupId = button.dataset.groupId;
        const section = document.querySelector(`[data-group-card="${CSS.escape(groupId)}"]`);
        const graded = section ? section.querySelectorAll('.question-card[data-graded="1"]').length : 0;
        const total = section ? section.querySelectorAll('.question-card').length : 0;
        button.classList.toggle('active', graded > 0 && graded < total);
        if (section && section.dataset.collapsed === '1') {
          button.classList.remove('active');
        }
      });
    }

    function submitAll() {
      document.querySelectorAll('.question-card [data-submit]').forEach(button => button.click());
    }

    function resetAll() {
      document.querySelectorAll('.question-card [data-reset]').forEach(button => button.click());
    }

    function expandAll() {
      document.querySelectorAll('.group-card').forEach(section => {
        section.dataset.collapsed = '0';
        const list = section.querySelector('.question-list');
        if (list) list.style.display = 'grid';
        const note = section.querySelector('.group-note');
        if (note) note.style.display = 'block';
        const toggle = section.querySelector('[data-toggle-group]');
        if (toggle) toggle.textContent = 'Collapse';
      });
    }

    function collapseAll() {
      document.querySelectorAll('.group-card').forEach(section => {
        section.dataset.collapsed = '1';
        const list = section.querySelector('.question-list');
        if (list) list.style.display = 'none';
        const note = section.querySelector('.group-note');
        if (note) note.style.display = 'none';
        const toggle = section.querySelector('[data-toggle-group]');
        if (toggle) toggle.textContent = 'Expand';
      });
      updateGroupState();
    }

    function clearHighlights() {
      document.querySelectorAll('.group-btn').forEach(button => button.classList.remove('active'));
    }

    function jumpFirstGroup() {
      const target = document.querySelector('[data-group-card]');
      if (target) target.scrollIntoView({behavior: 'smooth', block: 'start'});
    }

    document.getElementById('grade-all').addEventListener('click', submitAll);
    document.getElementById('reset-all').addEventListener('click', resetAll);
    document.getElementById('expand-all').addEventListener('click', expandAll);
    document.getElementById('collapse-all').addEventListener('click', collapseAll);
    document.getElementById('clear-highlight').addEventListener('click', clearHighlights);
    document.getElementById('jump-first').addEventListener('click', jumpFirstGroup);

    renderGroups();
    updateSummary();
    if (GROUPS.length === 0) {
      document.getElementById('group-area').innerHTML = '<div class="empty">No questions found.</div>';
    }
  </script>
</body>
</html>
"""
    hero_copy = (
        "Markdown sources are parsed as grouped practice sets with optional notes, numbered questions, and explicit answers. "
        "XLSX sources are parsed from the reviewed workbook layout and rendered in the same offline page. "
        "Submit one question, one group, or the full set without any network dependency."
    )
    return (
        template
        .replace("__TITLE__", "Offline Quiz HTML")
        .replace("__HERO_COPY__", hero_copy)
        .replace("__QUESTION_DATA__", _json_script(question_payload))
        .replace("__GROUP_DATA__", _json_script(group_payload))
        .replace("__APP_STATE_KEY__", json.dumps(APP_STATE_KEY, ensure_ascii=False))
    )


def _default_answer_path(output: Path) -> Path:
    if output.suffix:
        return output.with_suffix(DEFAULT_ANSWER_SUFFIX)
    return output.with_name(output.name + DEFAULT_ANSWER_SUFFIX)


def build_output(inputs: Sequence[Path], output: Path, answers: Path | None = None) -> tuple[list[Question], Path, Path | None]:
    questions = load_questions(inputs)
    if not questions:
        raise SystemExit("No supported markdown or XLSX files found.")
    output.write_text(render_html(questions), encoding="utf-8")
    answer_path = answers
    if answer_path is None and any(question.source_kind == "markdown" for question in questions):
        answer_path = _default_answer_path(output)
    if answer_path is not None:
        answer_path.write_text(render_answer_markdown(questions), encoding="utf-8")
    return questions, output, answer_path


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Build a self-contained offline quiz HTML file from markdown or XLSX sources.")
    parser.add_argument("--input", nargs="+", default=["."], help="Input files or directories")
    parser.add_argument("--output", default=DEFAULT_OUTPUT_NAME, help="Output HTML file path")
    parser.add_argument("--answers", default="", help="Optional markdown answer file path")
    parser.add_argument("--no-recursive", action="store_true", help="Do not search directories recursively")
    return parser


def main() -> int:
    args = build_parser().parse_args()
    inputs = _collect_inputs(args.input, recursive=not args.no_recursive)
    if not inputs:
        print("No supported markdown or XLSX files found.")
        return 1
    answer_path = Path(args.answers) if args.answers else None
    questions, output_path, answer_path = build_output(inputs, Path(args.output), answer_path)
    message = f"Built {output_path} with {len(questions)} questions"
    if answer_path is not None:
        message += f"; answer file {answer_path}"
    print(message)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
