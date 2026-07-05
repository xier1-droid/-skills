from __future__ import annotations

import argparse
import json
import random
from collections import defaultdict
from dataclasses import asdict, dataclass
from html import escape
from pathlib import Path
from typing import Iterable, Sequence

import openpyxl

SOURCE_FILE_HEADER = "\u6765\u6e90\u6587\u4ef6"
QUESTION_STEM_HEADER = "\u9898\u76ee"
QUESTION_TYPE_SINGLE = "\u5355\u9009\u9898"
QUESTION_TYPE_MULTI = "\u591a\u9009\u9898"
QUESTION_TYPE_JUDGEMENT = "\u5224\u65ad\u9898"
DEFAULT_OUTPUT_NAME = "\u79bb\u7ebf\u9898\u5e93\u5237\u9898.html"
APP_STATE_KEY = "offline_quiz_app_state_v1"


@dataclass(frozen=True)
class Question:
    uid: str
    workbook: str
    sheet: str
    row_index: int
    source_file: str
    question_number: str
    question_type: str
    stem: str
    options: dict[str, str]
    my_answer: str
    answer: str
    result: str
    analysis: str


def _normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\u3000", " ").strip()


def _normalize_type(value: object) -> str:
    text = _normalize_text(value)
    if "\u5224\u65ad" in text:
        return QUESTION_TYPE_JUDGEMENT
    if "\u591a\u9009" in text:
        return QUESTION_TYPE_MULTI
    return QUESTION_TYPE_SINGLE


def _normalize_answer(value: object) -> str:
    text = _normalize_text(value).upper()
    if not text:
        return ""
    if text in {"\u5bf9", "\u6b63\u786e", "TRUE", "T", "YES", "Y"}:
        return "A"
    if text in {"\u9519", "\u9519\u8bef", "FALSE", "F", "NO", "N"}:
        return "B"
    letters = [ch for ch in text if ch in "ABCDEF"]
    if not letters:
        return text
    ordered: list[str] = []
    for ch in letters:
        if ch not in ordered:
            ordered.append(ch)
    if len(ordered) > 1:
        return "".join(sorted(ordered))
    return ordered[0]


def _compare_answer(my_answer: str, answer: str) -> str:
    if not my_answer:
        return "\u672a\u7b54"
    return "\u6b63\u786e" if my_answer == answer else "\u9519\u8bef"


def _row_has_question(row: Sequence[object]) -> bool:
    return len(row) >= 14 and _normalize_text(row[1]) and _normalize_text(row[4])


def _question_uid(workbook: Path, sheet_name: str, row_index: int) -> str:
    return f"{workbook.name}::{sheet_name}::{row_index}"


def load_question_bank(paths: Iterable[Path | str]) -> list[Question]:
    questions: list[Question] = []
    for raw_path in paths:
        path = Path(raw_path)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header or len(header) < 15:
                continue
            if _normalize_text(header[1]) != SOURCE_FILE_HEADER or _normalize_text(header[4]) != QUESTION_STEM_HEADER:
                continue
            for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not _row_has_question(row):
                    continue
                options = {
                    label: _normalize_text(row[idx])
                    for idx, label in enumerate("ABCDEF", start=5)
                    if idx < len(row) and _normalize_text(row[idx])
                }
                my_answer = _normalize_answer(row[11] if len(row) > 11 else "")
                answer = _normalize_answer(row[12] if len(row) > 12 else "")
                qtype = _normalize_type(row[3])
                analysis = _normalize_text(row[14] if len(row) > 14 else "")
                if not analysis and len(row) > 13:
                    analysis = _normalize_text(row[13])
                questions.append(
                    Question(
                        uid=_question_uid(path, ws.title, row_index),
                        workbook=path.name,
                        sheet=ws.title,
                        row_index=row_index,
                        source_file=_normalize_text(row[1]),
                        question_number=_normalize_text(row[2]),
                        question_type=qtype,
                        stem=_normalize_text(row[4]),
                        options=options,
                        my_answer=my_answer,
                        answer=answer,
                        result=_compare_answer(my_answer, answer),
                        analysis=analysis,
                    )
                )
    return questions


def _shuffle_ids(question_ids: Sequence[str], seed: int) -> list[str]:
    ids = list(question_ids)
    random.Random(seed).shuffle(ids)
    return ids


def new_mode2_state(question_ids: Sequence[str], seed: int = 0) -> dict[str, object]:
    return {"seed": seed, "round": 0, "cursor": 0, "pool": _shuffle_ids(question_ids, seed)}


def draw_mode2_batch(question_ids: Sequence[str], state: dict[str, object], batch_size: int = 20) -> tuple[list[str], dict[str, object]]:
    ids = list(question_ids)
    pool = list(state.get("pool", []))
    cursor = int(state.get("cursor", 0))
    seed = int(state.get("seed", 0))
    round_index = int(state.get("round", 0))
    batch: list[str] = []
    while len(batch) < batch_size:
        if cursor >= len(pool):
            round_index += 1
            pool = _shuffle_ids(ids, seed + round_index)
            cursor = 0
        take = min(batch_size - len(batch), len(pool) - cursor)
        if take <= 0:
            break
        batch.extend(pool[cursor:cursor + take])
        cursor += take
    next_state = {"seed": seed, "round": round_index, "cursor": cursor, "pool": pool}
    return batch, next_state


def _group_questions(questions: Sequence[Question]) -> dict[str, list[Question]]:
    groups: dict[str, list[Question]] = defaultdict(list)
    for question in questions:
        groups[question.source_file].append(question)
    return dict(sorted(groups.items(), key=lambda item: item[0]))


def _json_script(value: object) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":")).replace("</", "<\\/")


def render_html(questions: Sequence[Question]) -> str:
    groups = _group_questions(questions)
    question_payload = [asdict(question) for question in questions]
    group_payload = {name: [question.uid for question in items] for name, items in groups.items()}
    template = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>__TITLE__</title>
  <style>
    :root {
      color-scheme: light;
      --bg: #f4efe6;
      --bg-soft: #fffaf1;
      --ink: #1f1b16;
      --muted: #6b6258;
      --brand: #c06a2b;
      --brand-2: #0f6b5b;
      --line: rgba(31, 27, 22, 0.12);
      --shadow: 0 18px 50px rgba(48, 28, 11, 0.12);
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      min-height: 100vh;
      color: var(--ink);
      background:
        radial-gradient(circle at top left, rgba(192, 106, 43, 0.12), transparent 28%),
        radial-gradient(circle at top right, rgba(15, 107, 91, 0.12), transparent 24%),
        linear-gradient(180deg, #f8f2e7 0%, #f0e5d3 100%);
      font-family: "Noto Serif SC", "Source Han Serif SC", "Songti SC", "SimSun", serif;
    }
    .shell { max-width: 1480px; margin: 0 auto; padding: 22px; }
    .hero { display: grid; grid-template-columns: 1.3fr 0.9fr; gap: 18px; margin-bottom: 18px; }
    .panel {
      background: rgba(255, 250, 241, 0.9);
      border: 1px solid var(--line);
      border-radius: 24px;
      box-shadow: var(--shadow);
      backdrop-filter: blur(8px);
    }
    .hero-main { padding: 24px; position: relative; overflow: hidden; }
    .hero-main::after {
      content: '';
      position: absolute;
      inset: auto -10% -40% auto;
      width: 320px;
      height: 320px;
      background: radial-gradient(circle, rgba(192, 106, 43, 0.18), transparent 65%);
      pointer-events: none;
    }
    .eyebrow { text-transform: uppercase; letter-spacing: .22em; color: var(--brand-2); font-size: 12px; font-weight: 700; }
    h1 { margin: 12px 0 10px; font-size: clamp(32px, 4vw, 56px); line-height: 1.02; }
    .hero-copy { max-width: 62ch; color: var(--muted); font-size: 16px; line-height: 1.75; }
    .stats { display: grid; grid-template-columns: repeat(3, minmax(0, 1fr)); gap: 12px; margin-top: 18px; }
    .stat { background: rgba(255, 255, 255, 0.72); border: 1px solid var(--line); border-radius: 18px; padding: 14px 16px; }
    .stat strong { display: block; font-size: 28px; margin-bottom: 4px; }
    .hero-side { padding: 20px; display: flex; flex-direction: column; gap: 14px; justify-content: space-between; }
    .mode-buttons { display: grid; grid-template-columns: repeat(2, minmax(0, 1fr)); gap: 12px; }
    .mode-btn, .action-btn, .group-btn {
      border: 0;
      border-radius: 16px;
      padding: 13px 16px;
      cursor: pointer;
      font: inherit;
      transition: transform .18s ease, box-shadow .18s ease, background .18s ease;
    }
    .mode-btn { background: linear-gradient(135deg, var(--brand), #ea9151); color: #fff; font-weight: 700; box-shadow: 0 12px 24px rgba(192, 106, 43, 0.22); }
    .mode-btn.secondary { background: linear-gradient(135deg, var(--brand-2), #2b8c7b); }
    .action-btn { background: #fff; border: 1px solid var(--line); color: var(--ink); }
    .mode-btn:hover, .action-btn:hover, .group-btn:hover { transform: translateY(-1px); }
    .layout { display: grid; grid-template-columns: 320px minmax(0, 1fr); gap: 18px; align-items: start; }
    .sidebar { padding: 18px; position: sticky; top: 18px; }
    .sidebar h2, .content h2 { margin: 0 0 12px; font-size: 22px; }
    .group-list { display: grid; gap: 10px; max-height: calc(100vh - 260px); overflow: auto; padding-right: 4px; }
    .group-btn { width: 100%; text-align: left; background: #fff; border: 1px solid var(--line); }
    .group-btn.active { border-color: rgba(15, 107, 91, 0.38); box-shadow: inset 0 0 0 1px rgba(15, 107, 91, 0.22); background: linear-gradient(135deg, rgba(15, 107, 91, 0.08), rgba(192, 106, 43, 0.06)); }
    .group-title { display: flex; justify-content: space-between; gap: 10px; font-size: 15px; font-weight: 700; }
    .group-meta { color: var(--muted); font-size: 13px; margin-top: 6px; line-height: 1.5; }
    .content { padding: 18px; }
    .toolbar { display: flex; flex-wrap: wrap; gap: 10px; margin-bottom: 14px; }
    .toolbar .primary { background: linear-gradient(135deg, var(--brand-2), #2b8c7b); color: #fff; border: 0; }
    .question-card { border: 1px solid var(--line); border-radius: 22px; padding: 20px; background: rgba(255,255,255,.82); box-shadow: 0 14px 32px rgba(40, 23, 8, 0.08); }
    .question-head { display: flex; justify-content: space-between; gap: 12px; align-items: start; margin-bottom: 10px; }
    .badge { display: inline-flex; align-items: center; gap: 6px; padding: 6px 10px; border-radius: 999px; background: rgba(15, 107, 91, 0.1); color: var(--brand-2); font-size: 13px; font-weight: 700; margin-right: 8px; margin-bottom: 8px; }
    .stem { font-size: 22px; line-height: 1.65; margin: 10px 0 16px; }
    .options { display: grid; gap: 10px; }
    .option { display: grid; grid-template-columns: auto auto 1fr; gap: 10px; align-items: start; padding: 12px 14px; border: 1px solid var(--line); border-radius: 16px; background: #fff; cursor: pointer; }
    .option input { margin-top: 4px; }
    .option-key { font-weight: 700; color: var(--brand); min-width: 18px; }
    .option-text { line-height: 1.65; }
    .status { margin-top: 14px; padding: 14px 16px; border-radius: 16px; background: rgba(31,27,22,.04); border: 1px dashed rgba(31,27,22,.18); color: var(--muted); line-height: 1.7; }
    .status.ok { background: rgba(15, 107, 91, 0.08); color: var(--brand-2); border-style: solid; }
    .status.bad { background: rgba(192, 106, 43, 0.08); color: #8a4c18; border-style: solid; }
    .question-actions { display: flex; flex-wrap: wrap; gap: 10px; margin-top: 14px; }
    .question-actions .primary { background: linear-gradient(135deg, var(--brand), #ea9151); color: #fff; border: 0; }
    .progress { font-size: 14px; color: var(--muted); margin-bottom: 12px; }
    .empty { padding: 28px; text-align: center; color: var(--muted); }
    @media (max-width: 1080px) {
      .hero, .layout { grid-template-columns: 1fr; }
      .sidebar { position: static; }
      .group-list { max-height: none; }
      .stats { grid-template-columns: 1fr; }
    }
  </style>
</head>
<body>
  <div class="shell">
    <section class="hero">
      <div class="panel hero-main">
        <div class="eyebrow">offline quiz studio</div>
        <h1>__TITLE__</h1>
        <p class="hero-copy">__HERO_COPY__</p>
        <div class="stats">
          <div class="stat"><strong id="stat-total">0</strong><span>__STAT_TOTAL__</span></div>
          <div class="stat"><strong id="stat-groups">0</strong><span>__STAT_GROUPS__</span></div>
          <div class="stat"><strong id="stat-pool">20</strong><span>__STAT_POOL__</span></div>
        </div>
      </div>
      <div class="panel hero-side">
        <div>
          <div class="mode-buttons">
            <button class="mode-btn" id="mode-source">__MODE1__</button>
            <button class="mode-btn secondary" id="mode-random">__MODE2__</button>
          </div>
        </div>
        <div class="group-meta">__SIDE_COPY__</div>
      </div>
    </section>

    <section class="layout">
      <aside class="panel sidebar">
        <h2>__GROUP_TITLE__</h2>
        <div id="group-list" class="group-list"></div>
      </aside>
      <main class="panel content">
        <div class="toolbar">
          <button class="action-btn primary" id="start-current">__START_CURRENT__</button>
          <button class="action-btn" id="next-question">__NEXT__</button>
          <button class="action-btn" id="prev-question">__PREV__</button>
          <button class="action-btn" id="reshuffle">__RESHUFFLE__</button>
          <button class="action-btn" id="clear-state">__CLEAR__</button>
        </div>
        <div class="progress" id="progress"></div>
        <div id="question-area"></div>
      </main>
    </section>
  </div>

  <script id="question-data" type="application/json">__QUESTION_DATA__</script>
  <script id="group-data" type="application/json">__GROUP_DATA__</script>
  <script>
    const QUESTIONS = JSON.parse(document.getElementById('question-data').textContent);
    const GROUPS = JSON.parse(document.getElementById('group-data').textContent);
    const APP_STATE_KEY = __APP_STATE_KEY__;
    const TYPE_MULTI = __TYPE_MULTI__;
    const TYPE_JUDGEMENT = __TYPE_JUDGEMENT__;
    const ALL_IDS = QUESTIONS.map(q => q.uid);
    const QUESTION_BY_ID = Object.fromEntries(QUESTIONS.map(q => [q.uid, q]));
    const GROUP_NAMES = Object.keys(GROUPS);
    const MODE_SOURCE = 'source';
    const MODE_RANDOM = 'random';

    function shuffle(list, seed) {
      const arr = [...list];
      let s = seed >>> 0;
      function rand() {
        s += 0x6D2B79F5;
        let t = Math.imul(s ^ s >>> 15, 1 | s);
        t ^= t + Math.imul(t ^ t >>> 7, 61 | t);
        return ((t ^ t >>> 14) >>> 0) / 4294967296;
      }
      for (let i = arr.length - 1; i > 0; i -= 1) {
        const j = Math.floor(rand() * (i + 1));
        [arr[i], arr[j]] = [arr[j], arr[i]];
      }
      return arr;
    }

    function seedMode2State() {
      const seed = Math.floor(Math.random() * 1e9);
      return { seed, round: 0, cursor: 0, pool: shuffle([...ALL_IDS], seed) };
    }

    function normalizeMode2State(state) {
      if (!state || typeof state.seed !== 'number' || typeof state.round !== 'number' || typeof state.cursor !== 'number' || !Array.isArray(state.pool)) {
        return seedMode2State();
      }
      return state;
    }

    function loadAppState() {
      try {
        const raw = localStorage.getItem(APP_STATE_KEY);
        if (raw) {
          const parsed = JSON.parse(raw);
          if (parsed && typeof parsed === 'object') {
            return parsed;
          }
        }
      } catch (err) {
        console.warn(err);
      }
      return { mode: MODE_SOURCE, group: GROUP_NAMES[0] || '', batch: [], index: 0, mode2: seedMode2State() };
    }

    let appState = loadAppState();
    let mode2State = normalizeMode2State(appState.mode2);
    let currentMode = appState.mode === MODE_RANDOM ? MODE_RANDOM : MODE_SOURCE;
    let currentGroup = GROUP_NAMES.includes(appState.group) ? appState.group : (GROUP_NAMES[0] || '');
    let currentBatch = Array.isArray(appState.batch) && appState.batch.every(id => QUESTION_BY_ID[id]) ? appState.batch.slice() : [];
    let currentIndex = Number.isInteger(appState.index) ? appState.index : 0;

    function saveAppState() {
      appState = { mode: currentMode, group: currentGroup, batch: currentBatch, index: currentIndex, mode2: mode2State };
      localStorage.setItem(APP_STATE_KEY, JSON.stringify(appState));
    }

    function drawMode2Batch(size = 20) {
      const batch = [];
      while (batch.length < size) {
        if (mode2State.cursor >= mode2State.pool.length) {
          mode2State.round += 1;
          mode2State.pool = shuffle([...ALL_IDS], mode2State.seed + mode2State.round);
          mode2State.cursor = 0;
        }
        const take = Math.min(size - batch.length, mode2State.pool.length - mode2State.cursor);
        if (take <= 0) break;
        batch.push(...mode2State.pool.slice(mode2State.cursor, mode2State.cursor + take));
        mode2State.cursor += take;
      }
      saveAppState();
      return batch;
    }

    function normalizeAnswer(answer) {
      if (!answer) return '';
      const text = String(answer).toUpperCase();
      if (['\u5bf9', '\u6b63\u786e', 'TRUE', 'T', 'YES', 'Y'].includes(text)) return 'A';
      if (['\u9519', '\u9519\u8bef', 'FALSE', 'F', 'NO', 'N'].includes(text)) return 'B';
      return Array.from(new Set(text.split('').filter(ch => 'ABCDEF'.includes(ch)))).sort().join('') || text;
    }

    function compareAnswer(selected, answer) {
      if (!selected) return '\u672a\u7b54';
      return normalizeAnswer(selected) === normalizeAnswer(answer) ? '\u6b63\u786e' : '\u9519\u8bef';
    }

    function getCurrentQuestions() {
      if (currentMode === MODE_RANDOM) {
        return currentBatch.map(id => QUESTION_BY_ID[id]).filter(Boolean);
      }
      if (currentGroup && GROUPS[currentGroup]) {
        return GROUPS[currentGroup].map(id => QUESTION_BY_ID[id]).filter(Boolean);
      }
      return QUESTIONS;
    }

    function renderGroupList() {
      const host = document.getElementById('group-list');
      host.innerHTML = '';
      GROUP_NAMES.forEach(name => {
        const ids = GROUPS[name];
        const button = document.createElement('button');
        button.className = 'group-btn' + (name === currentGroup && currentMode === MODE_SOURCE ? ' active' : '');
        button.innerHTML = `<div class="group-title"><span>${name}</span><span>${ids.length} 题</span></div><div class="group-meta">点击后按该来源文件顺序刷题</div>`;
        button.addEventListener('click', () => {
          currentMode = MODE_SOURCE;
          currentGroup = name;
          currentBatch = ids.slice();
          currentIndex = 0;
          saveAppState();
          renderAll();
        });
        host.appendChild(button);
      });
    }

    function renderQuestion(question, index, total) {
      const wrap = document.createElement('article');
      wrap.className = 'question-card';
      const inputName = `q-${question.uid.replace(/::/g, '-')}`;
      const optionType = question.question_type === TYPE_MULTI ? 'checkbox' : 'radio';
      const options = question.question_type === TYPE_JUDGEMENT ? [['A', '\u6b63\u786e'], ['B', '\u9519\u8bef']] : Object.entries(question.options);
      wrap.innerHTML = `
        <div class="question-head">
          <div>
            <span class="badge">${index + 1} / ${total}</span>
            <span class="badge">${question.question_type}</span>
            <span class="badge">${question.source_file}</span>
          </div>
          <div class="badge">${question.question_number || '\u65e0\u9898\u53f7'}</div>
        </div>
        <div class="stem">${escapeHtml(question.stem)}</div>
        <div class="options">
          ${options.map(([label, text]) => `
            <label class="option">
              <input type="${optionType}" name="${inputName}" value="${label}">
              <span class="option-key">${label}</span>
              <span class="option-text">${escapeHtml(text)}</span>
            </label>`).join('')}
        </div>
        <div class="question-actions">
          <button type="button" class="action-btn primary" data-submit>\u5224\u5b9a\u7ed3\u679c</button>
        </div>
        <div class="status" data-status>\u5148\u9009\u62e9\uff0c\u7136\u540e\u70b9\u6309\u94ae\u67e5\u770b\u5224\u5b9a\u7ed3\u679c\u3001\u6b63\u786e\u7b54\u6848\u548c\u89e3\u6790\u3002</div>
      `;
      const inputs = wrap.querySelectorAll('input');
      const submitButton = wrap.querySelector('[data-submit]');
      const status = wrap.querySelector('[data-status]');

      function getSelected() {
        return Array.from(inputs).filter(input => input.checked).map(input => input.value).join('');
      }

      function renderPendingStatus() {
        const selected = getSelected();
        const selectedText = selected ? `\u5df2\u9009\u62e9\uff1a${selected}` : '\u8fd8\u6ca1\u9009\u62e9\u4efb\u4f55\u9009\u9879';
        status.className = 'status';
        status.innerHTML = `${escapeHtml(selectedText)}<br>\u70b9\u51fb\u201c\u5224\u5b9a\u7ed3\u679c\u201d\u67e5\u770b\u6b63\u8bef\u3001\u6b63\u786e\u7b54\u6848\u548c\u89e3\u6790\u3002`;
      }

      function renderResult() {
        const selected = getSelected();
        const verdict = compareAnswer(selected, question.answer);
        const answerText = question.question_type === TYPE_JUDGEMENT ? (normalizeAnswer(question.answer) === 'A' ? '\u6b63\u786e' : '\u9519\u8bef') : (question.answer || '');
        status.className = 'status' + (verdict === '\u6b63\u786e' ? ' ok' : verdict === '\u9519\u8bef' ? ' bad' : '');
        status.innerHTML = `<strong>${verdict}</strong> · \u6b63\u786e\u7b54\u6848\uff1a<strong>${answerText}</strong><br>\u89e3\u6790\uff1a${escapeHtml(question.analysis || '\u6682\u65e0\u89e3\u6790')}`;
      }

      inputs.forEach(input => input.addEventListener('change', renderPendingStatus));
      submitButton.addEventListener('click', renderResult);
      renderPendingStatus();
      return wrap;
    }

    function escapeHtml(text) {
      return String(text)
        .replace(/&/g, '&amp;')
        .replace(/</g, '&lt;')
        .replace(/>/g, '&gt;')
        .replace(/"/g, '&quot;')
        .replace(/'/g, '&#39;');
    }

    function renderAll() {
      const questions = getCurrentQuestions();
      if (!questions.length) {
        currentIndex = 0;
      } else {
        currentIndex = Math.max(0, Math.min(currentIndex, questions.length - 1));
      }
      document.getElementById('stat-total').textContent = QUESTIONS.length;
      document.getElementById('stat-groups').textContent = GROUP_NAMES.length;
      document.getElementById('stat-pool').textContent = '20';
      renderGroupList();
      const progress = document.getElementById('progress');
      const host = document.getElementById('question-area');
      host.innerHTML = '';
      if (!questions.length) {
        progress.textContent = '\u6682\u65e0\u53ef\u5c55\u793a\u9898\u76ee\u3002';
        host.innerHTML = '<div class="empty">\u5f53\u524d\u6ca1\u6709\u53ef\u5237\u7684\u9898\u76ee\u3002</div>';
        saveAppState();
        return;
      }
      const question = questions[currentIndex];
      host.appendChild(renderQuestion(question, currentIndex, questions.length));
      progress.textContent = currentMode === MODE_RANDOM
        ? `\u6a21\u5f0f 2 \u00b7 \u5f53\u524d\u7b2c ${currentIndex + 1} \u9898 / ${questions.length} \u9898\uff0c\u5df2\u8fdb\u5165\u7b2c ${mode2State.round + 1} \u8f6e\u62bd\u9898\u3002`
        : `\u6a21\u5f0f 1 \u00b7 ${currentGroup}\uff0c\u5171 ${questions.length} \u9898\u3002`;
      saveAppState();
    }

    function startCurrent() {
      if (currentMode === MODE_RANDOM) {
        if (!currentBatch.length) {
          currentBatch = drawMode2Batch(20);
          currentIndex = 0;
        }
      } else {
        currentBatch = (GROUPS[currentGroup] || []).slice();
        currentIndex = 0;
      }
      renderAll();
    }

    function nextQuestion() {
      const questions = getCurrentQuestions();
      if (!questions.length) return;
      currentIndex = Math.min(currentIndex + 1, questions.length - 1);
      renderAll();
    }

    function prevQuestion() {
      const questions = getCurrentQuestions();
      if (!questions.length) return;
      currentIndex = Math.max(currentIndex - 1, 0);
      renderAll();
    }

    function switchToModeSource() {
      currentMode = MODE_SOURCE;
      currentGroup = GROUP_NAMES[0] || '';
      currentBatch = (GROUPS[currentGroup] || []).slice();
      currentIndex = 0;
      renderAll();
    }

    function switchToModeRandom() {
      currentMode = MODE_RANDOM;
      currentBatch = drawMode2Batch(20);
      currentIndex = 0;
      renderAll();
    }

    function reseedMode2() {
      mode2State = seedMode2State();
      currentMode = MODE_RANDOM;
      currentBatch = drawMode2Batch(20);
      currentIndex = 0;
      renderAll();
    }

    function clearState() {
      localStorage.removeItem(APP_STATE_KEY);
      appState = { mode: MODE_SOURCE, group: GROUP_NAMES[0] || '', batch: [], index: 0, mode2: seedMode2State() };
      mode2State = appState.mode2;
      currentMode = MODE_SOURCE;
      currentGroup = GROUP_NAMES[0] || '';
      currentBatch = (GROUPS[currentGroup] || []).slice();
      currentIndex = 0;
      renderAll();
    }

    document.getElementById('mode-source').addEventListener('click', switchToModeSource);
    document.getElementById('mode-random').addEventListener('click', switchToModeRandom);
    document.getElementById('start-current').addEventListener('click', startCurrent);
    document.getElementById('next-question').addEventListener('click', nextQuestion);
    document.getElementById('prev-question').addEventListener('click', prevQuestion);
    document.getElementById('reshuffle').addEventListener('click', reseedMode2);
    document.getElementById('clear-state').addEventListener('click', clearState);

    renderAll();
  </script>
</body>
</html>"""
    return (
        template
        .replace("__TITLE__", "\u79bb\u7ebf\u9898\u5e93\u5237\u9898")
        .replace("__HERO_COPY__", "\u6a21\u5f0f 1 \u6309 <code>\u6765\u6e90\u6587\u4ef6</code> \u5207\u9898\u7ec4\uff1b\u6a21\u5f0f 2 \u4ece\u5168\u90e8\u9898\u76ee\u91cc\u505a\u65e0\u653e\u56de\u5faa\u73af\u62bd\u6837\uff0c\u6bcf\u8f6e\u90fd\u5148\u628a\u5168\u5e93\u6d17\u4e00\u904d\uff0c\u518d\u7ee7\u7eed\u62bd 20 \u9898\u3002\u9875\u9762\u5b8c\u5168\u79bb\u7ebf\uff0c\u53ef\u76f4\u63a5\u53cc\u51fb\u6253\u5f00\u3002")
        .replace("__STAT_TOTAL__", "\u603b\u9898\u6570")
        .replace("__STAT_GROUPS__", "\u6765\u6e90\u6587\u4ef6\u7ec4")
        .replace("__STAT_POOL__", "\u6a21\u5f0f 2 \u62bd\u9898\u6570")
        .replace("__MODE1__", "\u6a21\u5f0f 1<br>\u6765\u6e90\u6587\u4ef6\u76ee\u5f55")
        .replace("__MODE2__", "\u6a21\u5f0f 2<br>\u968f\u673a 20 \u9898")
        .replace("__SIDE_COPY__", "\u5373\u65f6\u5224\u9898\u5df2\u5f00\u542f\u3002\u5355\u9009\u9898\u3001\u590d\u9009\u9898\u548c\u5224\u65ad\u9898\u90fd\u4f1a\u81ea\u52a8\u9002\u914d\u63a7\u4ef6\uff0c\u5e76\u5728\u4f5c\u7b54\u540e\u7acb\u5373\u663e\u793a\u7b54\u6848\u4e0e\u89e3\u6790\u3002")
        .replace("__GROUP_TITLE__", "\u9898\u7ec4\u76ee\u5f55")
        .replace("__START_CURRENT__", "\u5f00\u59cb\u5f53\u524d\u9898\u7ec4")
        .replace("__NEXT__", "\u4e0b\u4e00\u9898")
        .replace("__PREV__", "\u4e0a\u4e00\u9898")
        .replace("__RESHUFFLE__", "\u91cd\u65b0\u6d17\u724c")
        .replace("__CLEAR__", "\u6e05\u7a7a\u6a21\u5f0f 2 \u8fdb\u5ea6")
        .replace("__QUESTION_DATA__", _json_script(question_payload))
        .replace("__GROUP_DATA__", _json_script(group_payload))
        .replace("__APP_STATE_KEY__", json.dumps(APP_STATE_KEY, ensure_ascii=False))
        .replace("__TYPE_MULTI__", json.dumps(QUESTION_TYPE_MULTI, ensure_ascii=False))
        .replace("__TYPE_JUDGEMENT__", json.dumps(QUESTION_TYPE_JUDGEMENT, ensure_ascii=False))
    )


def _collect_input_xlsx(inputs: Iterable[str]) -> list[Path]:
    files: list[Path] = []
    for raw in inputs:
        path = Path(raw)
        if path.is_file() and path.suffix.lower() == ".xlsx":
            files.append(path)
        elif path.is_dir():
            files.extend(sorted(path.glob("*.xlsx")))
    return files


def build_output(paths: Iterable[Path | str], output: Path) -> tuple[list[Question], Path]:
    questions = load_question_bank(paths)
    output.write_text(render_html(questions), encoding="utf-8")
    return questions, output


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Build an offline quiz HTML file from XLSX question banks.")
    parser.add_argument("--input", nargs="+", default=["."], help="Input xlsx file(s) or directories")
    parser.add_argument("--output", default=DEFAULT_OUTPUT_NAME, help="Output HTML file path")
    return parser


def main() -> int:
    args = build_parser().parse_args()
    inputs = _collect_input_xlsx(args.input)
    if not inputs:
        print("No xlsx files found.")
        return 1
    questions, output_path = build_output(inputs, Path(args.output))
    print(f"Built {output_path} with {len(questions)} questions")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
